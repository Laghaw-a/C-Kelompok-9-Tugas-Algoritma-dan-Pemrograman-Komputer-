import pandas as pd
import os
from openpyxl import load_workbook

file_path = 'Kos_Mate_Data.xlsx'

# Fungsi untuk menyimpan data tanpa menghapus data lama
def simpan_data(sheet_name, data_baru):
    if os.path.exists(file_path):
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='overlay') as writer:
            try:
                data_lama = pd.read_excel(file_path, sheet_name=sheet_name)
                data = pd.concat([data_lama, data_baru], ignore_index=True)
            except ValueError:
                data = data_baru
            data.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(file_path) as writer:
            data_baru.to_excel(writer, sheet_name=sheet_name, index=False)

# Fungsi untuk menampilkan data dari Excel
def tampilkan_data(sheet_name):
    if os.path.exists(file_path):
        try:
            data = pd.read_excel(file_path, sheet_name=sheet_name)
            if data.empty:
                print(f"Ayo tambahkan data {sheet_name} terlebih dahulu!")
            else:
                print(f"\nData pada {sheet_name} :")
                print(data.to_string(index=False))
        except ValueError:
            print(f"Ayo tambahkan data {sheet_name} terlebih dahulu!")
    else:
        print(f"Ayo tambahkan data {sheet_name} terlebih dahulu!")


# Fungsi untuk menghapus data berdasarkan input pengguna
def hapus_data(sheet_name, kriteria):
    if os.path.exists(file_path):
        # Baca data yang ada
        workbook = load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            data = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Tampilkan data saat ini
            print(f"Data {sheet_name} Saat Ini:\n", data)
            
            # Hapus data berdasarkan kriteria
            data_hapus = data[data.apply(lambda row: all(row[k] == v for k, v in kriteria.items()), axis=1)]
            if not data_hapus.empty:
                data = data.drop(data_hapus.index)
                print("Data yang Dihapus:\n", data_hapus)
            else:
                print("Data tidak ditemukan berdasarkan kriteria yang diberikan.")
            
            # Tulis data yang sudah dihapus ke file
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            print(f"Data {sheet_name} tidak ditemukan.")
    else:
        print("File tidak ditemukan.")

#fungsi get_data        

file_path = 'Kos_Mate_Data.xlsx'

def get_data(sheet_name):
    """
    Mengambil data dari sheet tertentu di file Excel.
    
    Parameters:
        sheet_name (str): Nama sheet yang ingin diambil datanya.
        
    Returns:
        pd.DataFrame: DataFrame berisi data dari sheet tertentu, atau None jika file/sheet tidak ditemukan.
    """
    if os.path.exists(file_path):
        try:
            data = pd.read_excel(file_path, sheet_name=sheet_name)
            if not data.empty:
                return data
            else:
                return None
        except ValueError:
            return None
    else:
        return None

